#!/usr/bin/env python3
from __future__ import annotations

import csv
import re
from pathlib import Path

from project_paths import DATA_CLEAN, DATA_RAW, VERSION, ensure_dirs


def normalize_header(value: object) -> str:
    if value is None:
        return ""
    cleaned = str(value).strip().lower()
    cleaned = re.sub(r"\\s+", "_", cleaned)
    return cleaned


def iter_rows_csv(path: Path):
    with path.open(newline="", encoding="utf-8-sig") as fh:
        reader = csv.reader(fh)
        headers = next(reader)
        normalized = [normalize_header(h) for h in headers]
        for row in reader:
            yield dict(zip(normalized, row))


def iter_rows_xlsx(path: Path):
    try:
        import openpyxl
    except ImportError as exc:
        raise SystemExit(
            "openpyxl is required for .xlsx input. Install dependencies with: "
            "pip install -r requirements.txt"
        ) from exc

    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    normalized = [normalize_header(h) for h in headers]
    for row in rows:
        record = {}
        for key, value in zip(normalized, row):
            if value is None:
                record[key] = ""
            elif isinstance(value, str):
                record[key] = value.strip()
            else:
                record[key] = value
        yield record


def load_iso3_mapping(data_clean: Path) -> dict[str, str]:
    mapping_path = data_clean / "country_iso3_mapping.csv"
    if not mapping_path.exists():
        raise SystemExit(f"Missing {mapping_path}. Run 01_country_mapping.py first.")

    mapping: dict[str, str] = {}
    with mapping_path.open(newline="", encoding="utf-8") as fh:
        reader = csv.DictReader(fh)
        for row in reader:
            if row.get("source_type") == "WHO" and row.get("iso3"):
                mapping[row["name"]] = row["iso3"]
    return mapping


def clean_who_file(
    input_path: Path,
    output_path: Path,
    iso3_map: dict[str, str],
    region_override: str | None = None,
) -> int:
    output_fields = [
        "region_name",
        "location_name",
        "iso3",
        "sex_name",
        "year",
        "number_suicides_2021",
        "crude_suicide_rate_2021",
        "age_standardized_suicide_rate_2021",
        "data_quality",
        "income_group",
    ]

    row_count = 0
    iterator = iter_rows_xlsx(input_path) if input_path.suffix == ".xlsx" else iter_rows_csv(input_path)

    with output_path.open("w", newline="", encoding="utf-8") as fh:
        writer = csv.DictWriter(fh, fieldnames=output_fields)
        writer.writeheader()

        for row in iterator:
            country = row.get("country", "").strip()
            if not country:
                continue

            region = region_override or row.get("region", "").strip()
            record = {
                "region_name": region,
                "location_name": country,
                "iso3": iso3_map.get(country, ""),
                "sex_name": row.get("sex", "").strip(),
                "year": 2021,
                "number_suicides_2021": row.get("number_suicides_2021", ""),
                "crude_suicide_rate_2021": row.get("crude_suicide_rate_2021", ""),
                "age_standardized_suicide_rate_2021": row.get(
                    "age_standardized_suicide_rate_2021", ""
                ),
                "data_quality": row.get("data_quality", "").strip(),
                "income_group": row.get("income_group", "").strip(),
            }
            writer.writerow(record)
            row_count += 1

    return row_count


def main() -> None:
    ensure_dirs()

    iso3_map = load_iso3_mapping(DATA_CLEAN)

    xlsx_input = DATA_RAW / "who_global_master.xlsx"
    csv_input = DATA_RAW / "who_global_master.csv"
    if xlsx_input.exists():
        global_input = xlsx_input
    elif csv_input.exists():
        global_input = csv_input
    else:
        raise SystemExit("Missing WHO global input: who_global_master.xlsx or who_global_master.csv")

    global_output = DATA_CLEAN / "who_2021_clean.csv"
    global_rows = clean_who_file(global_input, global_output, iso3_map)
    print(f"[{VERSION}] Wrote {global_rows} rows to {global_output}")

    region_files = {
        "who_africa_region_full.csv": "Africa",
        "who_americas_region_full.csv": "Americas",
        "who_emro_region_full.csv": "EMRO",
        "who_europe_region_full.csv": "Europe",
        "who_searo_region_full.csv": "SEARO",
        "who_wpro_region_full.csv": "WPRO",
    }

    for filename, region in region_files.items():
        input_path = DATA_RAW / filename
        if not input_path.exists():
            continue
        output_name = filename.replace("_region_full", "_clean")
        output_path = DATA_CLEAN / output_name
        rows = clean_who_file(input_path, output_path, iso3_map, region_override=region)
        print(f"[{VERSION}] Wrote {rows} rows to {output_path}")


if __name__ == "__main__":
    main()
